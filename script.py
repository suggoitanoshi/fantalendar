#!/bin/env python

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

wb = openpyxl.load_workbook('ultah.xlsx')
sheet = wb.active

count = 0
start = 4

currentCell = sheet['A'+str(count+start)]
while currentCell.value != 2018:
  if count > sheet.max_row: break
  count += 1
  currentCell = sheet['A'+str(count+start)]

base_image = Image.open('template.jpg').convert('RGBA')
names_layer = Image.new('RGBA', base_image.size, (255, 255, 255, 0))
d = ImageDraw.Draw(names_layer)

maxwidth = 135
maxheight = 43
xpos = [15, 190, 364, 537, 706, 876, 1039]

for hari in range(1, 8):
  currenty = 315
  for row in range(start, start+count):
    name = sheet[get_column_letter(hari)+str(row)].value
    if name is None or name == '': continue
    fontsize = 30
    font = ImageFont.truetype('Dosis-Regular.ttf', fontsize)
    while font.getsize(name)[0] > maxwidth or font.getsize(name)[1] > maxheight:
      fontsize -= 1
      font = ImageFont.truetype('Dosis-Regular.ttf', fontsize)
    if fontsize < 16:
      tempname = name.split(' ')
      name = [tempname[0]]
      name.append(' '.join(tempname[1:]))
      fontsize = 1
      while font.getsize(name[-1])[0] < maxwidth and font.getsize(name[-1])[1]*len(name) < maxheight:
        fontsize +=1
        font = ImageFont.truetype('Dosis-Regular.ttf', fontsize)
    if type(name) == list:
      yheight = font.getsize(name[0])[1]
      l = 0
      for n in name:
        d.text((xpos[hari-1], currenty + l*yheight), n, font=font, fill=(0,0,0,255))
        l += 1
    else:
      d.text((xpos[hari-1], currenty), name, font=font, fill=(0,0,0,255))
    currenty += font.getsize(name[0])[1]

composite = Image.alpha_composite(base_image, names_layer)
composite.show()

composite.save('out.png')